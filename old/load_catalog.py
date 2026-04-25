#!/usr/bin/env python3
"""
Carga el catálogo completo desde la hoja PRODUCT del Excel maestro
hacia la base de datos SQLite de la app.

Estructura del Excel (headers en fila 2, datos fila 3+):
  FAMILIA | SUBFAMILIA | NOMBRE | CÓDIGO | NORMA | ANCHO MM | ESPESOR MM |
  LONGITUD MM | M2 PLACA | PESO KG M2 | PESO PLACA KG | UD PALE | M2 PALE |
  PESO PALE KG | UD VENTA | UD LOGISTICA | PVP ORIGEN EUR | UNIDAD PRECIO |
  DESCUENTO DISTRIBUIDOR | PRECIO ARIAS EUR | HS CODE | RENDIMIENTO_M2 |
  USO | ACTIVO | DISPO TARANCON | NOTAS
"""

import sqlite3
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl no instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / 'fassa_ops.db'

EXCEL_CANDIDATES = [
    BASE_DIR.parent / 'Arias_Group_Master-System_v1.xlsx',
    BASE_DIR.parent / 'Product AriasGroup V1.xlsx',
]

EXCEL_PATH = None
for candidate in EXCEL_CANDIDATES:
    if candidate.exists():
        EXCEL_PATH = candidate
        break

if EXCEL_PATH is None:
    print("❌ No se encontró el Excel maestro.")
    sys.exit(1)

print(f"📂 Excel: {EXCEL_PATH}")

# Backup defensivo antes de tocar la DB
import shutil
from datetime import datetime
backup_dir = BASE_DIR / 'backups'
backup_dir.mkdir(exist_ok=True)
ts = datetime.now().strftime('%Y%m%d_%H%M%S')
backup_path = backup_dir / f'fassa_ops_pre_catalog_{ts}.db'
shutil.copy2(DB_PATH, backup_path)
print(f"💾 Backup creado: {backup_path.name}")

# Leer + validar Excel ANTES de tocar la DB
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['PRODUCT']

parsed_rows = []
errors = []
for r in range(3, ws.max_row + 1):
    familia = ws.cell(row=r, column=1).value
    if not familia:
        continue

    sku = ws.cell(row=r, column=4).value
    name = ws.cell(row=r, column=3).value
    unit_sale = ws.cell(row=r, column=15).value
    unit_price = ws.cell(row=r, column=17).value
    units_pallet = ws.cell(row=r, column=12).value
    sqm_pallet = ws.cell(row=r, column=13).value

    if not sku or not name:
        continue

    try:
        unit = str(unit_sale).strip() if unit_sale else 'ud'
        price = float(unit_price) if unit_price else 0
        if isinstance(units_pallet, str) and units_pallet.startswith('='):
            up = None
        else:
            up = float(units_pallet) if units_pallet else None
        if isinstance(sqm_pallet, str) and sqm_pallet.startswith('='):
            sp = None
        else:
            sp = float(sqm_pallet) if sqm_pallet else None
        parsed_rows.append((str(sku), str(name), str(familia), unit, price, up, sp))
    except (ValueError, TypeError) as e:
        errors.append(f'Fila {r} (SKU {sku}): {e}')

if errors:
    print('❌ Errores de parseo — nada se ha escrito en la DB:')
    for e in errors[:10]:
        print(f'   · {e}')
    sys.exit(1)

if not parsed_rows:
    print('❌ 0 filas válidas en el Excel — abortando sin tocar la DB.')
    sys.exit(1)

print(f"✓ {len(parsed_rows)} productos validados. Aplicando cambios...")

# Transacción atómica: o se aplica todo, o nada
conn = sqlite3.connect(DB_PATH)
try:
    db = conn.cursor()
    db.execute('BEGIN')
    db.execute("DELETE FROM system_components")
    db.execute("DELETE FROM products")
    for row in parsed_rows:
        db.execute(
            """INSERT INTO products
            (sku, name, category, source_catalog, unit, unit_price_eur,
             units_per_pallet, sqm_per_pallet)
            VALUES (?, ?, ?, 'Gypsotech Abr2026', ?, ?, ?, ?)""",
            row
        )
    conn.commit()
    print(f"✅ {len(parsed_rows)} productos cargados")
    print(f"📂 DB: {DB_PATH}")
except Exception as e:
    conn.rollback()
    print(f"❌ Error durante la carga: {e}")
    print(f"   Rollback aplicado. Backup disponible en: {backup_path}")
    sys.exit(1)
finally:
    conn.close()
