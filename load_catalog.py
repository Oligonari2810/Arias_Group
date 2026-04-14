#!/usr/bin/env python3
"""
Carga el catálogo completo desde la hoja PRODUCT del Excel maestro
hacia la base de datos SQLite de la app.

Estructura del Excel (headers en fila 2, datos fila 3+):
  FAMILIA | SUBFAMILIA | NOMBRE | CÓDIGO | NORMA | ANCHO MM | ESPESOR MM |
  LONGITUD MM | M2 PLACA | PESO KG M2 | PESO PLACA KG | UD PALE | M2 PALE |
  PESO PALE KG | UD VENTA | UD LOGISTICA | PVP ORIGEN EUR | UNIDAD PRECIO |
  DESCUENTO DISTRIBUIDOR | PRECIO ARIAS EUR | HS CODE | RENDIMIENTO_M2 |
  USO | ACTIVO | DISPO TARANCON | NOTAS
"""

import sqlite3
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl no instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / 'fassa_ops.db'

EXCEL_CANDIDATES = [
    BASE_DIR.parent / 'Arias_Group_Master-System_v1.xlsx',
    BASE_DIR.parent / 'Product AriasGroup V1.xlsx',
]

EXCEL_PATH = None
for candidate in EXCEL_CANDIDATES:
    if candidate.exists():
        EXCEL_PATH = candidate
        break

if EXCEL_PATH is None:
    print("❌ No se encontró el Excel maestro.")
    sys.exit(1)

print(f"📂 Excel: {EXCEL_PATH}")

conn = sqlite3.connect(DB_PATH)
db = conn.cursor()
db.execute("DELETE FROM system_components")
db.execute("DELETE FROM products")
conn.commit()
print("🧹 Catálogo anterior eliminado.")

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['PRODUCT']

loaded = 0
for r in range(3, ws.max_row + 1):
    familia = ws.cell(row=r, column=1).value
    if not familia:
        continue
    
    sku = ws.cell(row=r, column=4).value
    name = ws.cell(row=r, column=3).value
    unit_sale = ws.cell(row=r, column=15).value
    unit_price = ws.cell(row=r, column=17).value
    units_pallet = ws.cell(row=r, column=12).value
    sqm_pallet = ws.cell(row=r, column=13).value
    
    if not sku or not name:
        continue
    
    unit = str(unit_sale).strip() if unit_sale else 'ud'
    price = float(unit_price) if unit_price else 0
    if isinstance(units_pallet, str) and units_pallet.startswith('='):
        up = None
    else:
        up = float(units_pallet) if units_pallet else None
    # Skip formula cells (start with =)
    if isinstance(sqm_pallet, str) and sqm_pallet.startswith('='):
        sp = None
    else:
        sp = float(sqm_pallet) if sqm_pallet else None
    
    db.execute(
        """INSERT OR REPLACE INTO products
        (sku, name, category, source_catalog, unit, unit_price_eur,
         units_per_pallet, sqm_per_pallet)
        VALUES (?, ?, ?, 'Gypsotech Abr2026', ?, ?, ?, ?)""",
        (str(sku), str(name), str(familia), unit, price, up, sp)
    )
    loaded += 1

conn.commit()
print(f"✅ {loaded} productos cargados")
print(f"📂 DB: {DB_PATH}")
conn.close()
